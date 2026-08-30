#!/usr/bin/env python3
"""
Perfexxxxion Pro Shop -- Excel -> Supabase migration.

Reads migration/source.xlsx and writes normalized CSVs to migration/output/:

    orders.csv          every order, deduped, normalized
    appointments.csv    every appointment, normalized
    items.csv           distinct item names for the autocomplete
    review_needed.csv   subset of the above that a human should eyeball
    report.txt          what happened, per sheet

Nothing is dropped. Rows with problems still land in orders.csv /
appointments.csv, tagged with migration_flag, and are ALSO copied into
review_needed.csv so they can be reviewed without holding up the import.

Usage:  python migrate.py
"""

import csv
import datetime as dt
import os
import re
import sys
import warnings
from collections import Counter, OrderedDict

warnings.filterwarnings("ignore")

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl is required:  python -m pip install openpyxl")

HERE = os.path.dirname(os.path.abspath(__file__))
SOURCE = os.path.join(HERE, "source.xlsx")
OUTDIR = os.path.join(HERE, "output")


# --------------------------------------------------------------------------
# Normalization
# --------------------------------------------------------------------------

# Every spelling of a location found in the workbook, mapped to canonical form.
LOCATION_MAP = {
    "valley": "Valley",
    "valley lanes": "Valley",
    "valley bowling lanes": "Valley",
    "carbondale": "Valley",            # confirmed: Carbondale == Valley
    "south": "South Side",
    "south side": "South Side",
    "southside": "South Side",
    "scranton": "South Side",          # confirmed: Scranton == Southside
    "idle hours south": "South Side",
    "idle hours": "South Side",
    "both": "Both",
}
# Values that appear in a location column but mean nothing.
LOCATION_JUNK = {"", "x", "-", "--", "---", "n/a", "na", "none", "."}

TRUE_VALUES = {"y", "yes", "true", "1", "paid", "pd"}
FALSE_VALUES = {"n", "no", "na", "n/a", "false", "0", "", "-", "--", "---", "."}

# Placeholder junk that shows up in name/id columns.
PLACEHOLDERS = {"", "-", "--", "---", ".", "n/a", "na", "(leave blank)", "none"}

LEGACY_ID_RE = re.compile(r"^\d{8}-[0-9A-Z]{3,4}$")


def is_blank(v):
    return v is None or str(v).strip() == ""


def text(v):
    """Cell -> clean string, or None."""
    if v is None:
        return None
    s = str(v).strip()
    if s.lower() in PLACEHOLDERS:
        return None
    return s or None


def norm_location(v):
    """Cell -> 'Valley' | 'South Side' | 'Both' | None. Returns (value, unmapped)."""
    if v is None:
        return None, None
    s = str(v).strip().lower()
    if s in LOCATION_JUNK:
        return None, None
    if s in LOCATION_MAP:
        return LOCATION_MAP[s], None
    # tolerate internal whitespace, e.g. "Valley  Lanes"
    collapsed = re.sub(r"\s+", " ", s)
    if collapsed in LOCATION_MAP:
        return LOCATION_MAP[collapsed], None
    return None, str(v).strip()          # unmapped -- caller flags it


def norm_bool(v):
    if v is None:
        return False
    s = str(v).strip().lower()
    if s in TRUE_VALUES:
        return True
    if s in FALSE_VALUES:
        return False
    return False


def norm_phone(v):
    """
    Cell -> ('570-555-1234' | None, leftover_text | None).

    Phones in the workbook are variously floats (5709779123.0), dashed
    strings, parenthesized, space separated, or plain words like
    'messenger'. Anything that isn't a usable number comes back as
    leftover text so the caller can push it into notes.
    """
    if v is None:
        return None, None
    if isinstance(v, float):
        v = int(v)                       # 5709779123.0 -> 5709779123
    s = str(v).strip()
    if s.lower() in PLACEHOLDERS:
        return None, None
    digits = re.sub(r"\D", "", s)
    if len(digits) == 11 and digits.startswith("1"):
        digits = digits[1:]
    if len(digits) == 10:
        return "%s-%s-%s" % (digits[:3], digits[3:6], digits[6:]), None
    if not digits:
        return None, s                   # e.g. 'messenger', a person's name
    return None, s                       # wrong digit count -- keep verbatim


def norm_date(v):
    """Cell -> (datetime.date | None, unparsed_text | None)."""
    if v is None:
        return None, None
    if isinstance(v, dt.datetime):
        return v.date(), None
    if isinstance(v, dt.date):
        return v, None
    s = str(v).strip()
    if s.lower() in PLACEHOLDERS:
        return None, None
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y", "%Y/%m/%d",
                "%m-%d-%Y", "%B %d, %Y", "%b %d, %Y"):
        try:
            return dt.datetime.strptime(s, fmt).date(), None
        except ValueError:
            pass
    return None, s


def norm_timestamp(v):
    if v is None:
        return None, None
    if isinstance(v, dt.datetime):
        return v, None
    if isinstance(v, dt.date):
        return dt.datetime.combine(v, dt.time()), None
    s = str(v).strip()
    if s.lower() in PLACEHOLDERS:
        return None, None
    for fmt in ("%Y-%m-%d %H:%M:%S.%f", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d",
                "%m/%d/%Y %H:%M:%S", "%m/%d/%Y %H:%M", "%m/%d/%Y"):
        try:
            return dt.datetime.strptime(s, fmt), None
        except ValueError:
            pass
    return None, s


def norm_time(v):
    if v is None:
        return None, None
    if isinstance(v, dt.time):
        return v, None
    if isinstance(v, dt.datetime):
        return v.time(), None
    s = str(v).strip()
    if s.lower() in PLACEHOLDERS:
        return None, None
    for fmt in ("%H:%M:%S", "%H:%M", "%I:%M %p", "%I:%M%p", "%I %p"):
        try:
            return dt.datetime.strptime(s.upper(), fmt).time(), None
        except ValueError:
            pass
    return None, s


def norm_price(v):
    if v is None:
        return None, None
    if isinstance(v, (int, float)):
        return round(float(v), 2), None
    s = str(v).strip().replace("$", "").replace(",", "")
    if s.lower() in PLACEHOLDERS:
        return None, None
    try:
        return round(float(s), 2), None
    except ValueError:
        return None, str(v).strip()


def norm_quantity(v):
    if v is None:
        return 1, None
    if isinstance(v, (int, float)):
        n = int(v)
        return (n, None) if n > 0 else (1, None)
    s = str(v).strip()
    if s.lower() in PLACEHOLDERS:
        return 1, None
    try:
        n = int(float(s))
        return (n, None) if n > 0 else (1, None)
    except ValueError:
        return 1, str(v).strip()


def norm_order_no(v):
    """Order # / Invoice # -- kill the float artifact (543601.0 -> 543601)."""
    if v is None:
        return None
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    s = str(v).strip()
    if s.lower() in PLACEHOLDERS:
        return None
    if s.endswith(".0") and s[:-2].isdigit():
        return s[:-2]
    return s


def norm_supplier(v):
    s = text(v)
    return re.sub(r"\s+", " ", s).strip() if s else None


# --------------------------------------------------------------------------
# Column layouts
#
# The workbook uses three generations of column order. Each layout maps a
# canonical field name to a zero-based column index.
# --------------------------------------------------------------------------

# 2026: raw Google Form headers. Used by Shop Order Ledger, 2026 Q1/Q2,
# Sheet26, and by the shifted lower half of All Orders.
LAYOUT_FORM_2026 = {
    "timestamp": 0, "email": 1, "customer_name": 2, "phone": 3,
    "order_location": 4, "pickup_location": 5, "item": 6, "fitting": 7,
    "notes": 8, "shop_order_date": 9, "supplier": 11, "supplier_order_no": 12,
    "price": 13, "invoice_no": 14, "paid": 15, "out_the_door": 16,
}

# 2025 Q4: adds Location, keeps Quantity and Shop Order Date.
LAYOUT_2025_Q4 = {
    "legacy_id": 0, "customer_name": 1, "phone": 2, "quantity": 3, "item": 4,
    "notes": 5, "pickup_location": 6, "submitted_date": 7, "order_due": 8,
    "staff_member": 9, "shop_order_date": 10, "supplier": 11,
    "supplier_order_no": 12, "price": 13, "invoice_no": 14,
    "internal_notes": 15, "paid": 16, "out_the_door": 17,
    "has_been_called": 18, "time_called": 19,
}

# 2025 Q3: same as Q4 but with no Location column.
LAYOUT_2025_Q3 = {
    "legacy_id": 0, "customer_name": 1, "phone": 2, "quantity": 3, "item": 4,
    "notes": 5, "submitted_date": 6, "order_due": 7, "staff_member": 8,
    "shop_order_date": 9, "supplier": 10, "supplier_order_no": 11,
    "price": 12, "invoice_no": 13, "internal_notes": 14, "paid": 15,
    "out_the_door": 16, "has_been_called": 18, "time_called": 19,
}

# 2025 Q1/Q2: oldest. No Location AND no Shop Order Date -- see build_order.
LAYOUT_2025_H1 = {
    "legacy_id": 0, "customer_name": 1, "phone": 2, "quantity": 3, "item": 4,
    "notes": 5, "submitted_date": 6, "order_due": 7, "staff_member": 8,
    "supplier": 9, "supplier_order_no": 10, "price": 11, "invoice_no": 12,
    "internal_notes": 13, "paid": 14, "out_the_door": 15,
    "has_been_called": 17, "time_called": 18,
}

# All Orders, upper half (rows 2..40): Dec 2025, its own layout.
LAYOUT_ALL_ORDERS_TOP = {
    "legacy_id": 0, "customer_name": 1, "phone": 2, "quantity": 3, "item": 4,
    "notes": 5, "pickup_location": 6, "submitted_date": 7, "order_due": 8,
    "staff_member": 9, "shop_order_date": 10, "supplier": 11,
    "supplier_order_no": 12, "price": 13, "invoice_no": 14,
    "internal_notes": 15, "paid": 16, "out_the_door": 17,
}

LAYOUT_NAMES = [
    ("2026 form layout", LAYOUT_FORM_2026),
    ("2025 Q4 layout", LAYOUT_2025_Q4),
    ("2025 Q3 layout", LAYOUT_2025_Q3),
    ("2025 H1 layout", LAYOUT_2025_H1),
    ("legacy layout", LAYOUT_ALL_ORDERS_TOP),
]

# Sheets to read, in dedupe-priority order. First sheet to claim a row wins,
# so the most complete / most current sheet is listed first.
ORDER_SHEETS = [
    ("Shop Order Ledger",           LAYOUT_FORM_2026, 1),
    ("Shop Order Listing 2026 Q2",  LAYOUT_FORM_2026, 1),
    ("Shop Order Listing 2026 Q1",  LAYOUT_FORM_2026, 1),
    ("Shop Order Listing 2025 Q4",  LAYOUT_2025_Q4,   1),
    ("Shop Order Listing 2025 Q3",  LAYOUT_2025_Q3,   1),
    ("Shop Order Listing 2025 Q2",  LAYOUT_2025_H1,   1),
    ("Shop Order Listing 2025 Q1",  LAYOUT_2025_H1,   1),
    ("All Orders",                  None,             1),   # dual layout
    ("Sheet26",                     LAYOUT_FORM_2026, 1),   # dup of 2026 Q2
]

# Sheets deliberately not migrated.
SKIPPED_SHEETS = OrderedDict([
    ("DashboardData", "derived view of All Orders, no unique data"),
    ("Service Tracker", "out of scope for v1 (dropped per spec)"),
    ("Sheet28", "read separately as the item autocomplete source"),
    ("Sheet29", "empty"),
    ("Time", "shop hours, not an order or appointment"),
    ("DONT TOUCH PAST THIS SHEET ", "intentionally blank"),
    ("Appointments", "read separately by read_appointments()"),
])


ORDER_FIELDS = [
    "legacy_id", "submitted_at", "source", "customer_name", "is_stock",
    "phone", "email", "item", "quantity", "fitting", "notes",
    "order_location", "pickup_location", "shop_order_date", "supplier",
    "supplier_order_no", "invoice_no", "price", "paid", "out_the_door",
    "has_been_called", "time_called", "staff_member", "internal_notes",
    "migration_flag", "source_sheet", "source_row",
]

APPT_FIELDS = [
    "legacy_id", "customer_name", "phone", "service", "location",
    "appt_date", "appt_time", "completed", "notes",
    "migration_flag", "source_sheet", "source_row",
]


def layout_label(layout):
    for name, obj in LAYOUT_NAMES:
        if layout is obj:
            return name
    return "unknown layout"


def looks_like_form_row(row):
    """
    All Orders is two sheets stacked. The lower portion was pasted in from the
    2026 Google Form export, which shifts every field two columns to the right
    -- customer names land in Phone Number, ball names land in Item Description.

    Detecting that by the '---' placeholders alone is not enough: plenty of the
    pasted rows carry a real timestamp in column A and a real email in column B.
    Two independent signals, either of which is conclusive:

      1. Column B holds an email address or the export's '---' placeholder.
         In the legacy layout column B is the customer's name.
      2. Columns E and F both normalize to known locations. In the legacy
         layout those columns are Item Description and Notes, which will not
         both read as 'Valley' or 'South Side'.
    """
    col_b = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ""
    if "@" in col_b or col_b == "---":
        return True

    loc_e, _ = norm_location(row[4] if len(row) > 4 else None)
    loc_f, _ = norm_location(row[5] if len(row) > 5 else None)
    return bool(loc_e and loc_f)


def cell(row, layout, field):
    idx = layout.get(field)
    if idx is None or idx >= len(row):
        return None
    return row[idx]


def build_order(row, layout, sheet, rownum, layout_name):
    """One spreadsheet row -> one canonical order dict, or None."""
    flags = []
    notes_extra = []

    name = text(cell(row, layout, "customer_name"))
    item = text(cell(row, layout, "item"))
    if not item:
        return None                      # no item == not an order row

    is_stock = bool(name and name.strip().lower() in ("stock", "shop", "shop stock"))
    if not name:
        name = "Unknown"
        flags.append("no customer name")

    phone, phone_leftover = norm_phone(cell(row, layout, "phone"))
    if phone_leftover:
        notes_extra.append("phone on file: %s" % phone_leftover)
        flags.append("unparseable phone")

    shop_order_date, sod_bad = norm_date(cell(row, layout, "shop_order_date"))
    submitted_at, ts_bad = norm_timestamp(cell(row, layout, "timestamp"))
    if submitted_at is None:
        sub_date, sub_bad = norm_date(cell(row, layout, "submitted_date"))
        if sub_date:
            submitted_at = dt.datetime.combine(sub_date, dt.time())
        elif sub_bad:
            flags.append("unparseable order date '%s'" % sub_bad)
    if ts_bad:
        flags.append("unparseable timestamp '%s'" % ts_bad)
    if sod_bad:
        flags.append("unparseable shop order date '%s'" % sod_bad)

    # The 2025 Q1/Q2 sheets predate the Shop Order Date column. Those orders
    # were all placed a year or more ago, so falling back to Order Date keeps
    # them out of the "not ordered yet" queue -- which is otherwise where all
    # ~500 of them would pile up, permanently, at the top of Amy's screen.
    if shop_order_date is None and "shop_order_date" not in layout:
        if submitted_at:
            shop_order_date = submitted_at.date()
            flags.append("shop order date inferred from order date")
        else:
            flags.append("no shop order date and no order date")

    # One row carries 0225-10-14, a slipped keystroke for 2025-10-14. Flagged
    # rather than auto-corrected -- guessing at a date is not our call.
    for label, d in (("shop order date", shop_order_date),
                     ("order date", submitted_at.date() if submitted_at else None)):
        if d and not (2000 <= d.year <= dt.date.today().year + 2):
            flags.append("%s looks mistyped (year %d)" % (label, d.year))

    order_loc, order_loc_bad = norm_location(cell(row, layout, "order_location"))
    pickup_loc, pickup_loc_bad = norm_location(cell(row, layout, "pickup_location"))
    for bad in (order_loc_bad, pickup_loc_bad):
        if bad:
            flags.append("unrecognized location '%s'" % bad)

    price, price_bad = norm_price(cell(row, layout, "price"))
    if price_bad:
        notes_extra.append("price as written: %s" % price_bad)
        flags.append("unparseable price '%s'" % price_bad)

    quantity, qty_bad = norm_quantity(cell(row, layout, "quantity"))
    if qty_bad:
        notes_extra.append("quantity as written: %s" % qty_bad)
        flags.append("unparseable quantity '%s'" % qty_bad)

    time_called, _ = norm_timestamp(cell(row, layout, "time_called"))

    legacy = text(cell(row, layout, "legacy_id"))
    if legacy and not LEGACY_ID_RE.match(legacy):
        legacy = None                    # timestamps and '---' are not IDs

    notes = text(cell(row, layout, "notes"))
    order_due = text(cell(row, layout, "order_due"))
    if order_due:
        notes_extra.append("order due: %s" % order_due)
    if notes_extra:
        notes = "; ".join([n for n in [notes] + notes_extra if n])

    has_ts = cell(row, layout, "timestamp") is not None
    source = "google_form" if (has_ts and submitted_at) else "import"

    return {
        "legacy_id": legacy or "",
        "submitted_at": submitted_at.isoformat(sep=" ") if submitted_at else "",
        "source": source,
        "customer_name": "Stock" if is_stock else name,
        "is_stock": is_stock,
        "phone": phone or "",
        "email": text(cell(row, layout, "email")) or "",
        "item": item,
        "quantity": quantity,
        "fitting": text(cell(row, layout, "fitting")) or "",
        "notes": notes or "",
        "order_location": order_loc or "",
        "pickup_location": pickup_loc or "",
        "shop_order_date": shop_order_date.isoformat() if shop_order_date else "",
        "supplier": norm_supplier(cell(row, layout, "supplier")) or "",
        "supplier_order_no": norm_order_no(cell(row, layout, "supplier_order_no")) or "",
        "invoice_no": norm_order_no(cell(row, layout, "invoice_no")) or "",
        "price": price if price is not None else "",
        "paid": norm_bool(cell(row, layout, "paid")),
        "out_the_door": norm_bool(cell(row, layout, "out_the_door")),
        "has_been_called": norm_bool(cell(row, layout, "has_been_called")),
        "time_called": time_called.isoformat(sep=" ") if time_called else "",
        "staff_member": text(cell(row, layout, "staff_member")) or "",
        "internal_notes": text(cell(row, layout, "internal_notes")) or "",
        "migration_flag": "; ".join(flags),
        "source_sheet": "%s (%s)" % (sheet, layout_name),
        "source_row": rownum,
    }


def dedupe_key(o):
    """
    Natural key for an order. Sheet26 duplicates 2026 Q2 wholesale, and the
    lower half of All Orders duplicates 2026 Q1, so matching on the business
    facts is more reliable than matching on any ID column.
    """
    return (
        (o["customer_name"] or "").lower(),
        re.sub(r"\s+", " ", (o["item"] or "").lower()),
        o["shop_order_date"] or o["submitted_at"][:10],
        str(o["price"]),
        (o["supplier_order_no"] or "").lower(),
    )


def read_orders(wb, report):
    seen = {}
    kept = []
    dupes = 0

    for sheet, layout, header_row in ORDER_SHEETS:
        if sheet not in wb.sheetnames:
            report.append("  !! sheet not found: %s" % sheet)
            continue
        ws = wb[sheet]
        n_read = n_kept = n_dupe = n_skip = 0

        for rownum, row in enumerate(
            ws.iter_rows(min_row=header_row + 1, values_only=True),
            start=header_row + 1,
        ):
            if all(is_blank(v) for v in row):
                continue
            n_read += 1

            if layout is None:
                if looks_like_form_row(row):
                    use, lname = LAYOUT_FORM_2026, "2026 form layout"
                else:
                    use, lname = LAYOUT_ALL_ORDERS_TOP, "legacy layout"
            else:
                use, lname = layout, layout_label(layout)

            rec = build_order(row, use, sheet, rownum, lname)
            if rec is None:
                n_skip += 1
                continue

            key = dedupe_key(rec)
            if key in seen:
                n_dupe += 1
                dupes += 1
                continue
            seen[key] = rec
            kept.append(rec)
            n_kept += 1

        report.append(
            "  %-30s read %4d | kept %4d | duplicate %4d | not an order %3d"
            % (sheet, n_read, n_kept, n_dupe, n_skip)
        )

    return kept, dupes


def read_appointments(wb, report):
    """Appointments sheet -- note the header is on row 2, not row 1."""
    ws = wb["Appointments"]
    out = []
    n_read = n_skip = 0

    for rownum, row in enumerate(ws.iter_rows(min_row=3, values_only=True), start=3):
        if all(is_blank(v) for v in row):
            continue
        n_read += 1
        flags = []

        legacy = text(row[0]) if len(row) > 0 else None
        name = text(row[1]) if len(row) > 1 else None
        service = text(row[3]) if len(row) > 3 else None

        # The sheet carries a filled-in template row ("(leave blank)" /
        # "Y or N") that is instructions, not an appointment.
        completed_raw = ""
        if len(row) > 8 and row[8] is not None:
            completed_raw = str(row[8]).strip().lower()
        if completed_raw == "y or n":
            n_skip += 1
            continue
        if not name and not service:
            n_skip += 1
            continue

        if legacy and not LEGACY_ID_RE.match(legacy):
            legacy = None

        phone, phone_leftover = norm_phone(row[2] if len(row) > 2 else None)
        notes = []
        if phone_leftover:
            notes.append("phone on file: %s" % phone_leftover)
            flags.append("unparseable phone")

        location, loc_bad = norm_location(row[4] if len(row) > 4 else None)
        if loc_bad:
            flags.append("unrecognized location '%s'" % loc_bad)

        appt_date, date_bad = norm_date(row[5] if len(row) > 5 else None)
        if date_bad:
            flags.append("unparseable date '%s'" % date_bad)
        if appt_date is None:
            flags.append("no appointment date")
        elif appt_date.year < 2024:
            # e.g. 2006-08-20 with ID 20060820-4IOV -- almost certainly a
            # mistyped 2026. Flagged, not silently corrected.
            flags.append("date looks mistyped (year %d)" % appt_date.year)

        appt_time, time_bad = norm_time(row[6] if len(row) > 6 else None)
        if time_bad:
            flags.append("unparseable time '%s'" % time_bad)

        submitted, _ = norm_timestamp(row[7] if len(row) > 7 else None)
        if submitted and appt_date and submitted.date() > appt_date:
            flags.append("submitted after the appointment date")

        if not service:
            service = "(not recorded)"
            flags.append("no service description")

        out.append({
            "legacy_id": legacy or "",
            "customer_name": name or "Unknown",
            "phone": phone or "",
            "service": service,
            "location": location or "",
            "appt_date": appt_date.isoformat() if appt_date else "",
            "appt_time": appt_time.strftime("%H:%M") if appt_time else "",
            "completed": norm_bool(row[8] if len(row) > 8 else None),
            "notes": "; ".join(notes),
            "migration_flag": "; ".join(flags),
            "source_sheet": "Appointments",
            "source_row": rownum,
        })

    report.append("  %-30s read %4d | kept %4d | template/blank %3d"
                  % ("Appointments", n_read, len(out), n_skip))
    return out


def read_items(wb, orders, report):
    """
    Sheet28 is 2,048 rows of item name + order date -- the raw material for
    the item autocomplete. Merged with item names from the migrated orders so
    nothing recent is missing.
    """
    counts = Counter()
    last_seen = {}

    if "Sheet28" in wb.sheetnames:
        for row in wb["Sheet28"].iter_rows(min_row=2, values_only=True):
            name = text(row[0] if len(row) > 0 else None)
            if not name:
                continue
            d, _ = norm_date(row[1] if len(row) > 1 else None)
            counts[name] += 1
            if d and (name not in last_seen or d > last_seen[name]):
                last_seen[name] = d

    for o in orders:
        name = o["item"]
        if not name:
            continue
        counts[name] += 1
        if o["shop_order_date"]:
            d = dt.date.fromisoformat(o["shop_order_date"])
            if name not in last_seen or d > last_seen[name]:
                last_seen[name] = d

    items = [{"item": n,
              "times_ordered": c,
              "last_ordered": last_seen[n].isoformat() if n in last_seen else ""}
             for n, c in counts.most_common()]
    report.append("  %-30s %d distinct item names" % ("Sheet28 + orders", len(items)))
    return items


def write_csv(path, fields, rows):
    with open(path, "w", newline="", encoding="utf-8") as fh:
        w = csv.DictWriter(fh, fieldnames=fields, extrasaction="ignore")
        w.writeheader()
        for r in rows:
            w.writerow(r)


def main():
    if not os.path.exists(SOURCE):
        sys.exit("missing %s" % SOURCE)
    os.makedirs(OUTDIR, exist_ok=True)

    wb = openpyxl.load_workbook(SOURCE, data_only=True)
    report = ["Perfexxxxion Pro Shop migration",
              "run: %s" % dt.datetime.now().strftime("%Y-%m-%d %H:%M"),
              "source: %s" % os.path.basename(SOURCE),
              "", "ORDERS"]

    orders, dupes = read_orders(wb, report)
    report.append("")
    report.append("APPOINTMENTS")
    appts = read_appointments(wb, report)
    report.append("")
    report.append("ITEMS")
    items = read_items(wb, orders, report)

    report.append("")
    report.append("SKIPPED SHEETS")
    for name, why in SKIPPED_SHEETS.items():
        if name in wb.sheetnames:
            report.append("  %-30s %s" % (name.strip(), why))

    unordered = [o for o in orders if not o["shop_order_date"]]
    flagged_orders = [o for o in orders if o["migration_flag"]]
    flagged_appts = [a for a in appts if a["migration_flag"]]

    report.append("")
    report.append("TOTALS")
    report.append("  orders kept                                       %d" % len(orders))
    report.append("  orders dropped as duplicates                      %d" % dupes)
    report.append("  orders with no shop order date (pin to top)       %d" % len(unordered))
    report.append("  appointments kept                                 %d" % len(appts))
    report.append("  distinct items                                    %d" % len(items))
    report.append("  flagged for review                                %d orders, %d appointments"
                  % (len(flagged_orders), len(flagged_appts)))

    report.append("")
    report.append("FLAG BREAKDOWN")
    fc = Counter()
    for r in flagged_orders + flagged_appts:
        for f in r["migration_flag"].split("; "):
            fc[re.sub(r"'[^']*'", "'...'", f)] += 1
    for f, c in fc.most_common():
        report.append("  %-52s %d" % (f, c))

    report.append("")
    report.append("VALUE CHECK (post-normalization)")
    for label, vals in (
        ("order_location", Counter(o["order_location"] or "(blank)" for o in orders)),
        ("pickup_location", Counter(o["pickup_location"] or "(blank)" for o in orders)),
        ("supplier", Counter(o["supplier"] or "(blank)" for o in orders)),
        ("appt location", Counter(a["location"] or "(blank)" for a in appts)),
    ):
        report.append("  %s: %s" % (label, dict(vals.most_common(10))))

    report.append("")
    report.append("DATE RANGE")
    dates = sorted(o["shop_order_date"] for o in orders if o["shop_order_date"])
    if dates:
        report.append("  orders:       %s .. %s" % (dates[0], dates[-1]))
    adates = sorted(a["appt_date"] for a in appts if a["appt_date"])
    if adates:
        report.append("  appointments: %s .. %s" % (adates[0], adates[-1]))

    write_csv(os.path.join(OUTDIR, "orders.csv"), ORDER_FIELDS, orders)
    write_csv(os.path.join(OUTDIR, "appointments.csv"), APPT_FIELDS, appts)
    write_csv(os.path.join(OUTDIR, "items.csv"),
              ["item", "times_ordered", "last_ordered"], items)

    review = [dict(r, record_type="order") for r in flagged_orders] + \
             [dict(r, record_type="appointment") for r in flagged_appts]
    write_csv(os.path.join(OUTDIR, "review_needed.csv"),
              ["record_type", "migration_flag", "source_sheet", "source_row",
               "customer_name", "item", "service", "phone", "shop_order_date",
               "appt_date", "appt_time", "price", "order_location",
               "pickup_location", "location", "notes"],
              review)

    text_report = "\n".join(report)
    with open(os.path.join(OUTDIR, "report.txt"), "w", encoding="utf-8") as fh:
        fh.write(text_report + "\n")
    print(text_report)


if __name__ == "__main__":
    main()
